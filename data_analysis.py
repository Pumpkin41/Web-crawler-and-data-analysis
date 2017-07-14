#-*-coding:utf-8-*-#
import re
import openpyxl
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt
#row行column列
'''print(sheet['A1'].value)
print(sheet['A1'].column)
print(sheet['A1'].row)'''

#根据姓名查找所在行号，将相应信息输出
def getrow(name):
    #print("the name is:",name)
    array=[]
    arr=0
    max_row=sheet.max_row+1
    for i in range(1,max_row,1):
        #print("the i is:",i)
        #print("the max_row is:",max_row)
        name2=name
        #text=r'(?<=name[:1])(?=name[1:])'
        content=sheet.cell(row=i,column=1).value
        #print(content)
        #text_list=re.match(text,content)
        #print(len(name))
        #print("aaaaaaaaa")
        if content is None:
            continue
        #print("len(content) is:",len(content))
        j=0
        if len(name) == len(content):
            for j in range(0,len(name),1):
                if name[j]==content[j]:
                    j=j+1
                else:
                    break
        #print("j is:",j)
        if j==len(name):
            #print("i is:",i)
            break
        else:
            continue
    return i
#根据行号获取合作者姓名
def getInfor(num):
    print('the name of Co-author are:')
    for s in range(num,num+5,1):
        print(sheet.cell(row=s,column=6).value)
#根据行号获取导师姓名
def getmanager(num):
    print("Read the blog manager is:")
    manager=sheet.cell(row=num,column=5).value
    print(manager)
    return manager
#根据行号获取合作者姓名，并将姓名赋值到数组namex[]
def getname(num):
    l=0
    scholar=sheet.cell(row=num,column=1).value
    namex= [ [ ' ' for i in range(0,20,1) ] for j in range(0,15,1)]
    '''namex=np.zeros((20,15))
    print(namex)
    namex=[([]*20) for i in range(15)]
    nameb=np.array(namex)
    print(nameb)'''
    for i in range(num,num+5,1):
        if l==0:
            pass
        else:
            l=l+1
        content=sheet.cell(row=i,column=6).value
        if content is None:
            continue
        k=0
        for j in range(0,len(content),1):
            #print("the k is:",k)
            #print("the l is:",l)
            #print("the j is:",j)
            #print("content[j] is:",content[j])
            #print("namex[l][k] is:",namex[l][k])
            namex[l][k]=content[j]
            k=k+1
            if content[j]==',':
                l=l+1
                k=0
                continue
            #print("namex[l] is:",namex[l])
    #for m in range(0,l+1,1):
        #print("namex[m] is:",namex[m])
    return namex
#将相应合作者信息以图形方式显示
def show_shape(name,namex):
    nameb=[ ' ' for i in range(0,15,1) ]
    '''for n in range(0,9,1):
        name_re[n]=input(namex[n]).slipt(",")
    print(name_re[0])'''
    #for j in range(0,14,1):
    nameb1=str(namex[0])
    #name1=str(namex[0])
    #print("nameb1:",nameb1)
    #for i in (0,14,1):
    nameb1="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[0][0],namex[0][1],namex[0][2],namex[0][3],namex[0][4],namex[0][5],namex[0][6],namex[0][7],namex[0][8],namex[0][9],namex[0][10],namex[0][11],namex[0][12],namex[0][13],namex[0][14],namex[0][15],namex[0][16],namex[0][17],namex[0][18],namex[0][19])
    nameb1=str(nameb1)
    
    nameb2=str(namex[0])
    nameb2="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[1][0],namex[1][1],namex[1][2],namex[1][3],namex[1][4],namex[1][5],namex[1][6],namex[1][7],namex[1][8],namex[1][9],namex[1][10],namex[1][11],namex[1][12],namex[1][13],namex[1][14],namex[1][15],namex[1][16],namex[1][17],namex[1][18],namex[1][19])
    nameb3=str(namex[0])
    nameb3="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[2][0],namex[2][1],namex[2][2],namex[2][3],namex[2][4],namex[2][5],namex[2][6],namex[2][7],namex[2][8],namex[2][9],namex[2][10],namex[2][11],namex[2][12],namex[2][13],namex[2][14],namex[2][15],namex[2][16],namex[2][17],namex[2][18],namex[2][19])
    nameb4=str(namex[0])
    nameb4="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[3][0],namex[3][1],namex[3][2],namex[3][3],namex[3][4],namex[3][5],namex[3][6],namex[3][7],namex[3][8],namex[3][9],namex[3][10],namex[3][11],namex[3][12],namex[3][13],namex[3][14],namex[3][15],namex[3][16],namex[3][17],namex[3][18],namex[3][19])
    nameb5=str(namex[0])
    nameb5="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[4][0],namex[4][1],namex[4][2],namex[4][3],namex[4][4],namex[4][5],namex[4][6],namex[4][7],namex[4][8],namex[4][9],namex[4][10],namex[4][11],namex[4][12],namex[4][13],namex[4][14],namex[4][15],namex[4][16],namex[4][17],namex[4][18],namex[4][19])
    nameb6=str(namex[0])
    nameb6="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[5][0],namex[5][1],namex[5][2],namex[5][3],namex[5][4],namex[5][5],namex[5][6],namex[5][7],namex[5][8],namex[5][9],namex[5][10],namex[5][11],namex[5][12],namex[5][13],namex[5][14],namex[5][15],namex[5][16],namex[5][17],namex[5][18],namex[5][19])
    nameb7=str(namex[0])
    nameb7="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[6][0],namex[6][1],namex[6][2],namex[6][3],namex[6][4],namex[6][5],namex[6][6],namex[6][7],namex[6][8],namex[6][9],namex[6][10],namex[6][11],namex[6][12],namex[6][13],namex[6][14],namex[6][15],namex[6][16],namex[6][17],namex[6][18],namex[6][19])
    nameb8=str(namex[0])
    nameb8="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[7][7],namex[7][1],namex[7][2],namex[7][3],namex[7][4],namex[7][5],namex[7][6],namex[7][7],namex[7][8],namex[7][9],namex[7][10],namex[7][11],namex[7][12],namex[7][13],namex[7][14],namex[7][15],namex[7][16],namex[7][17],namex[7][18],namex[7][19])
    nameb9=str(namex[0])
    nameb9="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[8][0],namex[8][1],namex[8][2],namex[8][3],namex[8][4],namex[8][5],namex[8][6],namex[8][7],namex[8][8],namex[8][9],namex[8][10],namex[8][11],namex[8][12],namex[8][13],namex[8][14],namex[8][15],namex[8][16],namex[8][17],namex[8][18],namex[8][19])
    nameb10=str(namex[0])
    nameb10="%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s%s"%(namex[9][0],namex[9][1],namex[9][2],namex[9][3],namex[9][4],namex[9][5],namex[9][6],namex[9][7],namex[9][8],namex[9][9],namex[9][10],namex[9][11],namex[9][12],namex[9][13],namex[9][14],namex[9][15],namex[9][16],namex[9][17],namex[9][18],namex[9][19])
    #print("nameb7 is:",nameb7)
    #print("nameb1 is:",nameb1)
    G=nx.Graph()
    G.add_node(name)
    G.add_nodes_from([name,nameb1,nameb2,nameb3,nameb4,nameb5,nameb6,nameb7,nameb8,nameb9])
    G.add_edges_from([(name,nameb1),(name,nameb2),(name,nameb3),(name,nameb4),(name,nameb5),(name,nameb6),(name,nameb7),(name,nameb8),(name,nameb9)])
    nx.draw_networkx(G,pos=None, arrows=True, with_labels=True)
    plt.title('relations of scholars')
    plt.show()

def show_manager_shape(name):
    num=getrow(name)
    managerx=getmanager(num)
    print("the manager name is:",managerx)
    num2=getrow(managerx)
    print("the num2 is:",num2)
    managerx2=getmanager(num2)
    num3=getrow(managerx2)
    managerx3=getmanager(num3)
    num4=getrow(managerx3)
    managerx4=getmanager(num4)
    G=nx.DiGraph()
    G.add_nodes_from([name,managerx,managerx2,managerx3,managerx4])
    G.add_edges_from([(name,managerx),(managerx,managerx2),(managerx2,managerx3),(managerx3,managerx4)])
    nx.draw_networkx(G,pos=None, arrows=True, with_labels=True)
    plt.title('relations of scholars')
    plt.show()

def getstudent(num):
    print("the name of student is:")
    student=[]
    for s in range(0,len(num),1):
        print("row number is:",num[s])
        student.append(sheet.cell(row=num[s],column=1).value)
    print("the name of student:",student)
    return student

def show_student_shape(name):
    stu_num_arr=[]
    stu_num=getrow(name)
    #max_stu_num=len(stu_num)
    student=getstudent(stu_num)
    #循环寻找下一层学生所在行号,行号保存在数组内
    max_student=len(student)
    for ss in range(0,max_student,1):
        #print("student[ss] is:",student[ss])
        if getrow(student[ss])==0:
            print(student[ss],"do not have student")
        else:
            stu_num_arr+=getrow(student[ss])
            print(student[ss],"'s student row number is:",getrow(student[ss]))
    '''for ii in (0,max_student,1):
        print("the stu_num_arr is:",stu_num_arr)'''
    print("2 ceng student row number:",stu_num_arr)
    if stu_num_arr==[]:
        G=nx.DiGraph()
        G.add_nodes_from([name,student[0],student[1]])
        G.add_edges_from([(name,student[0]),(name,student[1])])
        nx.draw_networkx(G,pos=None, arrows=True, with_labels=True)
        plt.title('relations of scholars')
        plt.show()
    else:
        student2=getstudent(stu_num_arr)
        G=nx.DiGraph()
        G.add_nodes_from([name,student[0],student[1]])
        G.add_edges_from([(name,student[0]),(name,student[1]),(student[1],student2[0])])
        nx.draw_networkx(G,pos=None, arrows=True, with_labels=True)
        plt.title('relations of scholars')
        plt.show()


#函数调用
data=openpyxl.load_workbook('informations.xlsx')
print(data.get_sheet_names())
sheet=data.get_sheet_by_name('Sheet2')
print('the sheet which is using:'+sheet.title)
name="Michael J. Frank"
#print(name[0])
num=getrow(name)
getInfor(num)
print("getInfor() is:",getInfor(num))
nameb=getname(num)
'''for i in range(0,10,1):
    print("the nameb is:",nameb[i])
    print(i)'''
#sheet=data.get_sheet_by_name('Sheet1')
#print('the sheet which is using:'+sheet.title)
#show_student_shape(name)
'''array2=getrow(name)
print("array2 is:",array2)
getstudent(array2)'''
#print("the num aa is:",num)
'''numx=[num]
max_num=len(numx)
print("max_num is:",max_num)
for i in range(0,max_num,1):
    print("the i xxx is:",i)
    managerx=getmanager(numx[i])
    numx.append(getrow(manager))'''
#show_manager_shape(name)
show_shape(name,nameb)