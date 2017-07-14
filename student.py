#-*-coding:utf-8-*-#
import re
import openpyxl
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt

#根据导师姓名查找所在行号，将相应信息输出，无返回0，有返回学生所在行号数组
def getrow(name):
    #print("the name is:",name)
    array=[]
    arr=0
    max_row=sheet.max_row+1
    for i in range(1,max_row,1):
        #print("the i is:",i)
        #print("the max_row is:",max_row)
        name2=name
        #text=r'(?<=name[:1])(?=name[1:])'
        content=sheet.cell(row=i,column=6).value
        #if content is not None:
        	#pass
            #print(content)
        #text_list=re.match(text,content)
        #print(len(name))
        #print("aaaaaaaaa")
        if content is None:
            continue
        #print("len(content) is:",len(content))
        j=0
        if len(name) == len(content):
            for j in range(0,len(name),1):
                if name[j]==content[j]:
                    j=j+1
                else:
                    break
        #print("j is:",j)
        if j==len(name):
            #print("i is:",i)
            array+=[i]
        else:
            continue
    #print("array is:",array)
    if array==[]:
        return 0
    else:
        return array

#根据行号获取学生姓名，返回学生姓名数组
def getstudent(num):
    #print("the name of student is:")
    student=[]
    for s in range(0,len(num),1):
        #print("row number is:",num[s])
        student.append(sheet.cell(row=num[s],column=1).value)
    #print("the name of student:",student)
    return student

#获取师生关系树
def getrelations(name):
    stu_num=[0]#行号数组，下表为0的数据为数组大小
    stu_name=[0]#姓名数组，下表为0的数据为数组大小
    stu_one=[2]#根据一个导师姓名获取的学生姓名数
    stu_name.append(name)#导师姓名存入数组，下标为1
    stu_name[0]+=1

    stu_num.append(getrow(name))#导师所在行号存入数组，下标为1
    print("getrow(name) is:",getrow(name))
    stu_num[0]+=1
    edge=[0]
    #获取导师的学生,最多查19层
    for i in range(1,20,1):
        #根据导师一个姓名查找出来的学生人数
        print("the number of student only once:",len(stu_name[i]))
        stu_one.append(len(stu_name[i]))
        
        #根据老师所在行号获取学生姓名
        print("the student names are:",getstudent(stu_num[i]))
        stu_name.append(getstudent(stu_num[i]))
        stu_name[0]+=1
        print(stu_name[i],"is the teacher of:",stu_name[i+1])
        #把新获得的学生当做老师，根据老师姓名获取下一代学生行号
        memory=0
        for j in range(0,len(stu_name[i+1]),1):
            if getrow(stu_name[i+1][j])==0:
                print(stu_name[i+1][j],"has no student!")
            else:
                stu_num.append(getrow(stu_name[i+1][j]))
                stu_num[0]+=1
                memory+=1
                #print("j is:",j)
                print("stu_name[i+1][j] is:",stu_name[i+1][j])
                edge.append(j)
        if memory==0:
            break
    print("stu_name is:",stu_name)
    print("edge is:",edge)
    #print("stu_name[2][0]:",stu_name[2][0])
    G=nx.DiGraph()
    #画点
    for m in range(1,len(stu_name),1):
        if len(stu_name[m])>5:
            #print("the m is:",x)
            G.add_node(stu_name[m])
        else:
            for n in range(0,len(stu_name[m]),1):
                G.add_node(stu_name[m][n])
    #G.add_nodes_from([stu_name[1],stu_name[2][0]])
    #画箭头
    for x in range(2,len(stu_name),1):
        if len(stu_name[x-1])>5:
            if(len(stu_name[x])<=5):
                for a in range(0,len(stu_name[x]),1):
                    G.add_edge(stu_name[x][a],stu_name[x-1])
        else:
            for y in range(0,len(stu_name[x]),1):
                G.add_edge(stu_name[x][y],stu_name[x-1][edge[1]])
    #G.add_edges_from([(stu_name[1],stu_name[2][0])])
    nx.draw_networkx(G,pos=None, arrows=True, with_labels=True)
    plt.title('relations of scholars')
    plt.show()

#
data=openpyxl.load_workbook('tutor.xlsx')
#print(data.get_sheet_names())
sheet=data.get_sheet_by_name('bbb')
print('the sheet which is using:'+sheet.title)
name='Andrew Barto'
#test=getrow(name)
#print("*******************",test)
getrelations(name)